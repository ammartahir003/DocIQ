import json
import groq
import fitz
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO


client = groq.Groq(api_key=settings.GROQ_API_KEY)

def extract_text(file_path: str) -> str:
    doc = fitz.open(file_path)
    text = ""
    for page in doc:
        text += page.get_text()
    return text

def summarize_pdf(file_path: str) -> str:
    text = extract_text(file_path)

    response = client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[
            {
                "role": "system",
                "content": "You are a document analysis assistant. Summarize documents clearly and concisely."
            },
            {
                "role": "user",
                "content": f"Summarize this document in 3 bullet points:\n\n{text[:8000]}"
            }
        ],
        max_tokens=1024
    )
    return response.choices[0].message.content

def answer_question(file_path: str, question: str, history: list) -> str:
    text = extract_text(file_path)

    messages = [
        {
            "role": "system",
            "content": f"You are a document assistant. Answer questions using only this document:\n\n{text[:8000]}"
        }
    ]

    for msg in history:
        messages.append({"role": msg["role"], "content": msg["content"]})

    messages.append({"role": "user", "content": question})

    response = client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=messages,
        max_tokens=1024
    )
    return response.choices[0].message.content

import json

def extract_structured_data(file_path: str) -> dict:
    text = extract_text(file_path)

    system_prompt = """You are a document data extraction engine. Analyze the document and respond with ONLY valid JSON, no markdown, no explanation, no code fences.

First detect the document type, then extract relevant fields using this schema:

{
  "document_type": "invoice" | "contract" | "resume" | "report" | "other",
  "fields": {
    // For invoice: vendor, invoice_number, date, due_date, total_amount, currency, line_items (array of {description, quantity, price})
    // For contract: parties (array), effective_date, expiry_date, key_terms (array of strings), obligations (array of strings)
    // For resume: name, email, phone, skills (array), experience (array of {company, role, duration}), education (array of {institution, degree, year})
    // For report/other: title, author, date, key_points (array of strings), entities_mentioned (array)
  }
}

Only include fields you can confidently find in the document. Use null for missing fields. Respond with raw JSON only."""

    response = client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": text[:8000]}
        ],
        max_tokens=1500,
        temperature=0
    )

    raw = response.choices[0].message.content.strip()
    raw = raw.replace("```json", "").replace("```", "").strip()

    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        return {"document_type": "other", "fields": {}, "error": "Could not parse structured data"}
    
def export_to_excel(documents):
    """documents: list of Document objects with extracted_data"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    # Collect all unique simple field keys across documents
    all_keys = []
    for doc in documents:
        if doc.extracted_data and doc.extracted_data.get('fields'):
            for key, value in doc.extracted_data['fields'].items():
                if not isinstance(value, list) and key not in all_keys:
                    all_keys.append(key)

    headers = ['Filename', 'Document Type'] + [k.replace('_', ' ').title() for k in all_keys]
    ws.append(headers)

    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for doc in documents:
        row = [doc.filename]
        if doc.extracted_data:
            row.append(doc.extracted_data.get('document_type', ''))
            fields = doc.extracted_data.get('fields', {})
            for key in all_keys:
                value = fields.get(key, '')
                if isinstance(value, list):
                    value = ''
                row.append(value if value is not None else '')
        else:
            row.append('')
            row.extend([''] * len(all_keys))
        ws.append(row)

    for col in ws.columns:
        max_len = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def search_across_documents(documents, question: str) -> dict:
    """
    documents: queryset of Document objects belonging to the user
    Returns: { answer: str, sources: [{id, filename}] }
    """
    # Build a compact index: filename + summary + key extracted fields per doc
    doc_index = []
    for doc in documents:
        snippet = f"[Document ID: {doc.id}] Filename: {doc.filename}\n"
        if doc.summary:
            snippet += f"Summary: {doc.summary}\n"
        if doc.extracted_data and doc.extracted_data.get('fields'):
            snippet += f"Extracted fields: {json.dumps(doc.extracted_data['fields'])}\n"
        doc_index.append(snippet)

    combined_index = "\n---\n".join(doc_index)

    system_prompt = """You are a cross-document search assistant. You are given summaries and extracted data from MULTIPLE documents. 

Answer the user's question using ONLY the information provided. If the answer involves specific documents, mention them by filename.

At the end of your answer, on a new line, write exactly: SOURCES: [list of Document IDs that contain relevant info, comma separated, e.g. SOURCES: 3,7,12]
If no documents are relevant, write SOURCES: none"""

    response = client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": f"DOCUMENTS:\n{combined_index[:10000]}\n\nQUESTION: {question}"}
        ],
        max_tokens=1024,
        temperature=0
    )

    raw = response.choices[0].message.content
    
    # Extract sources line
    sources = []
    if "SOURCES:" in raw:
        answer_part, sources_part = raw.rsplit("SOURCES:", 1)
        answer = answer_part.strip()
        ids_str = sources_part.strip()
        if ids_str.lower() != "none":
            try:
                sources = [int(i.strip()) for i in ids_str.split(",") if i.strip().isdigit()]
            except:
                sources = []
    else:
        answer = raw.strip()

    return {"answer": answer, "source_ids": sources}